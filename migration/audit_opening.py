"""Reconstruct the 2009 opening balance from the per-type sheets.

Sums every flight on each aircraft-type sheet, split by era (before the logbook
opens on 2009-01-08 vs after), so we can see which types/sim make up the
opening 'brought forward' balance recorded on page 1 of the LOG BOOK.
"""

import json
from datetime import datetime
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent
wb = openpyxl.load_workbook(HERE / "_tmp_read.xlsx", data_only=True)

# Military layout: flight total is column 9 ("Flight"); sim is col 14, IF actual col 15.
MIL = {"cols_total": 9, "sim": 14, "if": 15}
# Civilian layout: flying = cols 3..10; instrument IN FLIGHT col 11, SIM col 12.
CIV = {"fly": range(3, 11), "if": 11, "sim": 12}

SHEETS = {
    "B206": MIL, "CT4": MIL, "S70": MIL,
    "S92": CIV, "AW139": CIV, "R22": CIV, "FW SINGLES": CIV,
}
SIM_SHEETS = {"S70 FF&MS": MIL, "S92 SIM": CIV, "AW139 SIM": CIV}

CUTOFF = datetime(2009, 1, 8)


def parse_my(v, last):
    if v is None:
        return last
    s = str(v).strip()
    for fmt in ("%b %y", "%b-%y", "%B %y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return last


def num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def scan(name, spec, sim=False):
    ws = wb[name]
    my = None
    pre = post = sim_pre = sim_post = 0.0
    dmin = dmax = None
    for r in range(4, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        my = parse_my(a, my)
        # a flight row has a day number
        try:
            day = int(float(str(b).strip()))
        except (TypeError, ValueError):
            continue
        if my is None:
            continue
        dt = my.replace(day=min(day, 28))
        if spec is MIL:
            fly = num(ws.cell(row=r, column=9).value)
        else:
            fly = sum(num(ws.cell(row=r, column=c).value) for c in spec["fly"])
        if fly <= 0:
            continue
        if my < CUTOFF:
            pre += fly
        else:
            post += fly
        dmin = dt if dmin is None or dt < dmin else dmin
        dmax = dt if dmax is None or dt > dmax else dmax
    return round(pre, 1), round(post, 1), dmin, dmax


print(f"{'sheet':12} {'PRE-2009':>9} {'POST':>8}   date range")
print("-" * 50)
tot_pre = 0.0
for name, spec in SHEETS.items():
    pre, post, dmin, dmax = scan(name, spec)
    tot_pre += pre
    rng = f"{dmin:%b %Y} .. {dmax:%b %Y}" if dmin else ""
    print(f"{name:12} {pre:9.1f} {post:8.1f}   {rng}")
print("-" * 50)
print(f"{'TOTAL PRE-2009 flying (actual sheets)':40} {tot_pre:.1f}")

print("\nSIM sheets (flying-column basis):")
sim_pre_tot = 0.0
for name, spec in SIM_SHEETS.items():
    pre, post, dmin, dmax = scan(name, spec, sim=True)
    sim_pre_tot += pre
    rng = f"{dmin:%b %Y} .. {dmax:%b %Y}" if dmin else ""
    print(f"  {name:12} pre={pre:6.1f}  post={post:6.1f}   {rng}")
print(f"  --> pre-2009 sim total: {round(sim_pre_tot,1)}")

# Opening balance recorded on page 1
rep = json.loads((HERE / "reconciliation_report.json").read_text(encoding="utf-8"))
ob = rep["opening_balance"]
fly = sum(v for k, v in ob.items() if k not in ("inst_in_flight", "inst_sim"))
print(f"\nRecorded page-1 opening balance: flying={round(fly,1)}  "
      f"inst_in_flight={ob['inst_in_flight']}  inst_sim={ob['inst_sim']}")
