"""
Enrichment pass for the migrated logbook.

Takes flights.json (produced by migrate.py) and:
  * strips the aircraft REGO out of each note's crew line into its own field,
  * assigns the aircraft TYPE using a rego->type map built from the workbook's
    own per-aircraft-type sheets (AW139, S92, B206, ...),
  * derives Pilot-in-Command vs Other-crew from the flight type
    (Command time > 0  =>  SELF is PIC; otherwise the named person is PIC),
  * writes enriched_flights.json,
  * writes rego_review.csv listing any flight whose rego/type could not be
    confidently resolved, for a quick human check.
"""

from __future__ import annotations

import csv
import json
import re
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent
_SRC_XLSX = HERE / "Log Book - Hildebrand.xlsx"
_TMP_XLSX = HERE / "_tmp_read.xlsx"


def _resolve_xlsx() -> Path:
    """Read from a fresh temp copy so this works while the workbook is open in
    Excel (Windows locks the original for direct reads)."""
    import shutil

    try:
        shutil.copyfile(_SRC_XLSX, _TMP_XLSX)
        return _TMP_XLSX
    except OSError:
        return _TMP_XLSX if _TMP_XLSX.exists() else _SRC_XLSX


XLSX = _resolve_xlsx()
IN_FLIGHTS = HERE / "flights.json"
OUT_FLIGHTS = HERE / "enriched_flights.json"
OUT_REVIEW = HERE / "rego_review.csv"

OWNER = "SELF"

# Manual aircraft-type assignments for LOG BOOK rows the auto-detector can't
# type (no rego, note is just a route). Keyed by the flight's Excel source_row.
# These are genuine actual flights that belong on a per-type sheet.
MANUAL_TYPE: dict[int, str] = {
    104: "S92",    # 2009-12-14 Moresette, WBGR-OFFSHORE-WBGR (offshore S-92)
    105: "S92",    # 2009-12-16 SINAUL, WBGR-NAGA1-KK-OCRS-WBGR (offshore S-92)
    974: "S92",    # 2016-07-31 GASPAR NYZ, YTST-YMVT-YTST (S-92)
    1831: "AW139", # 2026-02-19 YSTW-BURRAPINE-YPMQ RNP21 (recent AW139)
}

# Per-aircraft-type sheets -> canonical type label. (SIM sheets excluded here;
# sim time is already captured in the inst_sim column.)
TYPE_SHEETS = {
    "AW139": "AW139",
    "S70": "S-70",
    "S70 FF&MS": "S-70",
    "B206": "B206",
    "R22": "R22",
    "FW SINGLES": "FW SINGLE",
    "CT4": "CT4",
    "S92": "S92",
}

FOREIGN_RE = re.compile(r"^[A-Za-z]-[A-Za-z0-9]{2,4}$")   # B-MHI, I-EPIC, M-MHL
AUS_RE = re.compile(r"^[A-Z]{3}$")                         # LOH, NYZ, XIB
MODEL_RE = re.compile(r"^[A-Z]{1,3}\d{2,3}[A-Z]?$")       # R22, DA20, C172, PA28, S76

# 3-letter tokens that are words / abbreviations / crew names, never a rego.
STOPWORDS = {
    "SELF", "AND", "THE", "DEM", "DEMO", "NEW", "OLD", "SIM", "IRT", "IPC",
    "IFR", "VFR", "NDB", "VOR", "DME", "ILS", "ARA", "SID", "CCT", "PIC",
    "FSI", "LOE", "OPC", "SEE", "FOR", "NIL", "TAA", "REC", "REF", "UK",
    "YEO",  # crewman name, not a rego
}


def rego_candidates(text: str) -> list[str]:
    """Return rego-like tokens found in a piece of text."""
    out = []
    for tok in re.split(r"[\s/]+", text or ""):
        tok = tok.strip()
        if not tok:
            continue
        if FOREIGN_RE.match(tok):
            out.append(tok.upper())
        elif AUS_RE.match(tok) and tok not in STOPWORDS:
            out.append(tok)
    return out


def norm_note(text: str) -> str:
    """Normalise a note for matching across sheets."""
    return re.sub(r"\s+", " ", (text or "").strip().upper())


def build_maps() -> tuple[dict[str, str], dict[str, str]]:
    """Scan the per-type sheets' comments to learn:

    * rego  -> type  (from rego tokens in the notes)
    * note  -> type  (from the full note text, so flights without a rego can
                      still be typed by matching the identical note)
    """
    wb = openpyxl.load_workbook(XLSX, data_only=False)
    token_types: dict[str, Counter] = defaultdict(Counter)
    note_types: dict[str, Counter] = defaultdict(Counter)
    for sheet, label in TYPE_SHEETS.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if cell.comment and cell.comment.text.strip():
                    text = cell.comment.text.strip()
                    for r in rego_candidates(text.splitlines()[0]):
                        token_types[r][label] += 1
                    note_types[norm_note(text)][label] += 1
    rego_type = {tok: t.most_common(1)[0][0] for tok, t in token_types.items()}
    # Only keep unambiguous note->type mappings (a note that maps to one type).
    note_type = {
        n: t.most_common(1)[0][0] for n, t in note_types.items() if len(t) == 1
    }
    return rego_type, note_type


def build_frequency_regos(flights: list[dict], min_count: int = 3) -> set[str]:
    """3-letter tokens that recur across crew lines are treated as regos."""
    counts: Counter = Counter()
    for f in flights:
        for tok in re.split(r"[\s/]+", f.get("crew", "")):
            tok = tok.strip()
            if AUS_RE.match(tok) and tok not in STOPWORDS:
                counts[tok] += 1
    return {tok for tok, n in counts.items() if n >= min_count}


def command_hours(times: dict) -> float:
    return (
        times["se_command_day"]
        + times["se_command_night"]
        + times["me_command_day"]
        + times["me_command_night"]
    )


def main() -> None:
    flights = json.loads(IN_FLIGHTS.read_text(encoding="utf-8"))

    rego_type, note_type = build_maps()
    freq_regos = build_frequency_regos(flights)
    known_regos = set(rego_type) | freq_regos

    review_rows = []
    stats = Counter()

    for f in flights:
        crew = f.get("crew", "")
        tokens = [t for t in re.split(r"\s+", crew) if t]

        rego = ""
        confidence = ""
        # 1) overseas hyphenated rego (unambiguous)
        for t in tokens:
            if FOREIGN_RE.match(t):
                rego, confidence = t.upper(), "foreign"
                break
        # 2) a token confirmed as a rego by the type sheets
        if not rego:
            for t in tokens:
                if t in rego_type:
                    rego, confidence = t, "type-sheet"
                    break
        # 3) a token that recurs often across crew lines
        if not rego:
            for t in tokens:
                if t in freq_regos:
                    rego, confidence = t, "frequency"
                    break

        # Aircraft type from the rego, else the matching note in a type sheet,
        # else an explicit model token in the note.
        aircraft_type = rego_type.get(rego, "")
        type_source = "rego" if aircraft_type else ""
        if not aircraft_type:
            aircraft_type = note_type.get(norm_note(f["note_raw"]), "")
            if aircraft_type:
                type_source = "note-match"
        model_token = next((t for t in tokens if MODEL_RE.match(t)), "")
        if not aircraft_type and model_token:
            aircraft_type, type_source = model_token, "model-token"

        # Manual override for hand-verified rows the heuristics can't type.
        if f["source_row"] in MANUAL_TYPE:
            aircraft_type, type_source = MANUAL_TYPE[f["source_row"]], "manual"

        # Clean crew name = tokens minus rego and model designators.
        drop = {rego} | ({model_token} if model_token else set())
        name = " ".join(t for t in tokens if t not in drop).strip()

        # Derive PIC / other crew.
        #   Command flight  -> you are PIC, the named person is other crew.
        #   Otherwise       -> the named person is PIC, and you (SELF) are the
        #                      other crew member.
        if command_hours(f["times"]) > 0:
            pic, other = OWNER, name
        else:
            pic, other = name, OWNER

        f["aircraftRego"] = rego
        f["aircraftType"] = aircraft_type
        f["pilotInCommand"] = pic
        f["otherCrew"] = other
        f["regoConfidence"] = confidence or "none"
        f["typeSource"] = type_source or "none"

        stats[confidence or "none"] += 1
        # Only rows we genuinely can't type need a human look.
        if not aircraft_type:
            review_rows.append(
                {
                    "source_row": f["source_row"],
                    "date": f["date"] or f"{f['year']}-{f['month']:02d}-??",
                    "note_raw": f["note_raw"],
                    "proposed_rego": rego,
                    "proposed_type": aircraft_type,
                    "clean_name": name,
                    "confidence": confidence or "none",
                }
            )

    OUT_FLIGHTS.write_text(json.dumps(flights, indent=2), encoding="utf-8")
    with OUT_REVIEW.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(
            fh,
            fieldnames=[
                "source_row", "date", "note_raw", "proposed_rego",
                "proposed_type", "clean_name", "confidence",
            ],
        )
        w.writeheader()
        w.writerows(review_rows)

    # ---- Console summary -------------------------------------------------
    print(f"Flights processed        : {len(flights)}")
    print(f"Distinct regos learned    : {len(known_regos)}")
    print(f"  from type sheets        : {len(rego_type)}")
    print(f"  from frequency          : {len(freq_regos)}")
    print()
    print("Rego match confidence:")
    for k in ("foreign", "type-sheet", "frequency", "none"):
        print(f"  {k:12s}: {stats.get(k, 0)}")
    typed = sum(1 for f in flights if f["aircraftType"])
    print()
    print("Type source:")
    for k in ("rego", "note-match", "model-token", "none"):
        print(f"  {k:12s}: {sum(1 for f in flights if f['typeSource'] == k)}")
    print()
    print(f"Flights with a rego       : {sum(1 for f in flights if f['aircraftRego'])}")
    print(f"Flights with a type       : {typed}")
    print(f"Rows needing review       : {len(review_rows)}  -> {OUT_REVIEW.name}")

    type_counts = Counter(f["aircraftType"] for f in flights if f["aircraftType"])
    print("\nType distribution:")
    for t, n in type_counts.most_common():
        print(f"  {t:10s}: {n}")

    print(f"\nWrote {OUT_FLIGHTS.name} and {OUT_REVIEW.name}")


if __name__ == "__main__":
    main()
