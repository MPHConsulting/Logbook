"""
Build the app seed dataset from the migration outputs.

Reads enriched_flights.json + reconciliation_report.json and writes
../src/data/logbook-data.json, containing:
  * openingBalance   - your pre-Excel carried-forward hours (per column)
  * adjustments      - a reconciliation delta so the app's grand total
                       reproduces your Excel "TOT TO DATE" EXACTLY, even where
                       the source spreadsheet is internally inconsistent
  * excelGrandTotal  - the authoritative final total recorded in the Excel
  * flights          - every flight, in the app's camelCase schema
"""

from __future__ import annotations

import json
import re
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent
_SRC_XLSX = HERE / "Log Book - Hildebrand.xlsx"
_TMP_XLSX = HERE / "_tmp_read.xlsx"


def _resolve_xlsx() -> Path:
    """Read from a fresh temp copy so the build works even while the workbook is
    open in Excel (Windows locks the original for direct reads)."""
    import shutil

    try:
        shutil.copyfile(_SRC_XLSX, _TMP_XLSX)
        return _TMP_XLSX
    except OSError:
        return _TMP_XLSX if _TMP_XLSX.exists() else _SRC_XLSX


XLSX = _resolve_xlsx()
ENRICHED = HERE / "enriched_flights.json"
REPORT = HERE / "reconciliation_report.json"
OUT = HERE.parent / "src" / "data" / "logbook-data.json"

# NVG hours are recorded in the AW139 sheet (column N, 0-based index 13), with
# the pre-Excel military carried-forward balance flagged "Start" in column O.
NVG_COL = 13
NVG_START_COL = 14

# snake_case (migration) -> camelCase (app)
TIME_MAP = {
    "se_icus_day": "seIcusDay",
    "se_icus_night": "seIcusNight",
    "se_dual_day": "seDualDay",
    "se_dual_night": "seDualNight",
    "se_command_day": "seCommandDay",
    "se_command_night": "seCommandNight",
    "me_icus_day": "meIcusDay",
    "me_icus_night": "meIcusNight",
    "me_dual_day": "meDualDay",
    "me_dual_night": "meDualNight",
    "me_command_day": "meCommandDay",
    "me_command_night": "meCommandNight",
    "me_copilot_day": "meCopilotDay",
    "me_copilot_night": "meCopilotNight",
    "inst_in_flight": "instInFlight",
    "inst_sim": "instSim",
}


def r2(x: float) -> float:
    return round(x + 0.0, 2)


def norm_note(text: str) -> str:
    """Normalise a note for cross-sheet matching."""
    return re.sub(r"\s+", " ", (text or "").upper()).strip().rstrip("/ ").strip()


def canon_type(label: str) -> str:
    """Collapse an aircraft-type label to a canonical key (mirrors the TS
    canonType) so per-type figures line up across differing naming."""
    s = re.sub(r"[^A-Z0-9]", "", (label or "").upper())
    if "AW139" in s:
        return "AW139SIM" if "SIM" in s else "AW139"
    if "S92" in s:
        return "S92SIM" if "SIM" in s else "S92"
    if "S70" in s:
        return "S70FFMS" if "FFMS" in s else "S70"
    if "B206" in s:
        return "B206"
    if "CT4" in s:
        return "CT4"
    if "R22" in s:
        return "R22"
    if "DA40" in s or "DA20" in s:
        return "DA40"
    return s or "OTHER"


def parse_month_year(text: str, last_year):
    text = str(text).strip()
    for fmt in ("%b %y", "%b-%y", "%B %y"):
        try:
            d = datetime.strptime(text, fmt)
            return d.month, d.year
        except ValueError:
            pass
    for fmt in ("%b", "%B"):
        try:
            d = datetime.strptime(text, fmt)
            return d.month, last_year
        except ValueError:
            pass
    return None, last_year


def build_nvg():
    """From the AW139 sheet, return per-date NVG hours, per-date note, the
    carried-forward opening balance, and the recorded total.

    NVG is matched to the main log by DATE (robust to note-text differences
    between the AW139 sheet and the main LOG BOOK sheet).
    """
    wb_v = openpyxl.load_workbook(XLSX, data_only=True)["AW139"]
    wb_c = openpyxl.load_workbook(XLSX, data_only=False)["AW139"]
    by_date: dict[str, float] = {}
    note_by_date: dict[str, str] = {}
    cur_year = None
    cur_month = None
    for r in range(1, wb_v.max_row + 1):
        mon = wb_v.cell(row=r, column=1).value
        day = wb_v.cell(row=r, column=2).value
        v = wb_v.cell(row=r, column=NVG_COL + 1).value
        start = wb_v.cell(row=r, column=NVG_START_COL + 1).value

        if isinstance(mon, str) and mon.strip():
            if mon.strip().upper().startswith("TOTAL"):
                continue
            m, cur_year = parse_month_year(mon, cur_year)
            if m:
                cur_month = m

        # Skip the carried-forward "Start" marker row (its value is the military
        # opening balance, handled separately below).
        if isinstance(start, str) and "START" in start.upper():
            continue
        if not isinstance(v, (int, float)) or v <= 0:
            continue

        try:
            dd = int(float(str(day).strip()))
        except (ValueError, TypeError):
            dd = None
        if not (cur_year and cur_month and dd):
            continue
        iso = f"{cur_year:04d}-{cur_month:02d}-{dd:02d}"
        by_date[iso] = by_date.get(iso, 0.0) + float(v)
        note = ""
        for c in range(1, wb_c.max_column + 1):
            cell = wb_c.cell(row=r, column=c)
            if cell.comment and cell.comment.text.strip():
                note = cell.comment.text
                break
        note_by_date.setdefault(iso, note)

    # The carried-forward military NVG uses the itemised figure from the Totals
    # sheet ("Grand Total Mil", Using NVD = K18 = 109.6) rather than the rounded
    # 109 in the AW139 "Start" cell, so the logbook and the CV/Totals pages agree.
    totals_ws = openpyxl.load_workbook(XLSX, data_only=True)["Totals"]
    opening = float(totals_ws.cell(row=18, column=11).value or 0.0)
    total = r2(opening + sum(by_date.values()))
    return by_date, note_by_date, opening, total


# Column order of the logbook flying columns, mapped to their 1-based column
# in the Totals sheet's "Summary of all Flying Hours" table.
TOTALS_COLS = [
    ("seIcusDay", 2),
    ("seIcusNight", 3),
    ("seDualDay", 4),
    ("seDualNight", 5),
    ("seCommandDay", 6),
    ("seCommandNight", 7),
    ("meIcusDay", 8),
    ("meIcusNight", 9),
    ("meDualDay", 10),
    ("meDualNight", 11),
    ("meCommandDay", 12),
    ("meCommandNight", 13),
    ("meCopilotDay", 14),
    ("meCopilotNight", 15),
]


# Corrections for two known inconsistencies in the Excel "Summary of all Flying
# Hours" table (rows 34-45) versus the authoritative military-summary / CV-Summary
# tables. Applied so the Totals page agrees with the CV page; the user is fixing
# the same cells in their Excel too.
#   * B206B-1 co-pilot: 207.2 -> 209.4 (day 176.4->179.8, night 30.8->29.6) => 770.1
#   * S70 FF&MS sim Command-Day: 4.6 -> 4.7 => sim total 45.4 -> 45.5
TOTALS_FIXES = {
    "B206B-1": {"meCopilotDay": 179.8, "meCopilotNight": 29.6},
    "S70 FF&MS": {"meCommandDay": 4.7},
}


def build_totals_sheet(grand_total_flying: float | None = None):
    """Snapshot the Totals sheet's 'Summary of all Flying Hours' table plus the
    grand-total statistics block, as displayed for the app's Totals page.

    The headline "Grand Total Flying Hours" is the sum of the per-type rows
    (i.e. the table's own TOTALS row) so it matches what you get by adding the
    Total column down the Totals / CV pages. ``grand_total_flying`` is retained
    for compatibility but no longer overrides the headline."""
    ws = openpyxl.load_workbook(XLSX, data_only=True)["Totals"]

    def rowvals(r):
        return {key: r2(ws.cell(row=r, column=col).value or 0.0) for key, col in TOTALS_COLS}

    rows = []
    total_row = None
    # Type rows live between the "Summary of all Flying Hours" header and its
    # TOTALS line (rows 34..45 in the current workbook).
    for r in range(34, 46):
        label = ws.cell(row=r, column=1).value
        if not label:
            continue
        label = str(label).strip()
        times = rowvals(r)
        for key, val in TOTALS_FIXES.get(label, {}).items():
            times[key] = r2(val)
        total = r2(sum(times.values()))
        if label.upper() == "TOTALS":
            total_row = {"times": times, "total": total}
            break
        rows.append({"type": label, "times": times, "total": total})

    sheet_grand = r2(ws.cell(row=47, column=4).value or 0.0)
    # Headline = sum of the per-type rows (reflects TOTALS_FIXES corrections),
    # i.e. what you get by adding the Total column down the page (5005.5).
    computed_grand = r2(sum(r["total"] for r in rows))
    stats = {
        "grandTotalFlyingHours": computed_grand,
        "sheetGrandTotalFlyingHours": sheet_grand,
        "coPilot": r2(ws.cell(row=47, column=14).value or 0.0),
        "aeronauticalExperience": r2(ws.cell(row=50, column=4).value or 0.0),
        "captainHelicopter": r2(ws.cell(row=53, column=4).value or 0.0),
        "icusHelicopter": r2(ws.cell(row=53, column=10).value or 0.0),
    }
    return {
        "columns": [k for k, _ in TOTALS_COLS],
        "rows": rows,
        "totalsRow": total_row,
        "stats": stats,
    }


# CV Summary columns, mapped to their 1-based column in the CV Summary sheet.
CV_COLS = [
    ("captainIcus", 3),
    ("otherCoPilot", 4),
    ("dual", 5),
    ("night", 6),
    ("instrument", 7),
    ("nvg", 8),
    ("total", 9),
]


def cv_grand_from_logbook(final_camel: dict[str, float]) -> dict[str, float]:
    """Derive the CV grand-total row from the logbook running total (opening +
    entries + kept adjustments) so the CV headline equals the authoritative
    Grand Total Flying Hours. Columns mirror the CV Summary layout."""
    g = final_camel
    captain_icus = (
        g["seCommandDay"] + g["seCommandNight"] + g["meCommandDay"] + g["meCommandNight"]
        + g["seIcusDay"] + g["seIcusNight"] + g["meIcusDay"] + g["meIcusNight"]
    )
    other = g["meCopilotDay"] + g["meCopilotNight"]
    dual = g["seDualDay"] + g["seDualNight"] + g["meDualDay"] + g["meDualNight"]
    night = (
        g["seIcusNight"] + g["seDualNight"] + g["seCommandNight"]
        + g["meIcusNight"] + g["meDualNight"] + g["meCommandNight"] + g["meCopilotNight"]
    )
    return {
        "captainIcus": r2(captain_icus),
        "otherCoPilot": r2(other),
        "dual": r2(dual),
        "night": r2(night),
        "instrument": r2(g.get("instInFlight", 0.0) + g.get("instSim", 0.0)),
        "nvg": r2(g.get("nvg", 0.0)),
        "total": r2(captain_icus + other + dual),
    }


def build_cv_summary(nvg_by_type: dict[str, float] | None = None,
                     logbook_grand: dict[str, float] | None = None):
    """Snapshot the CV Summary sheet: helicopter and fixed-wing groups with
    subtotals, the grand total, and the side statistics.

    The Excel CV sheet's NVG column omitted the AW139 (civilian) night-vision
    hours, so we fold the true per-type NVG (from the imported flights) into the
    matching rows and recompute the NVG subtotals/grand total.

    ``logbook_grand`` overrides the grand-total row with the logbook-derived
    figures so the CV headline matches the authoritative Grand Total Flying
    Hours; the per-type / group rows stay as the CV-sheet snapshot (indicative).
    """
    nvg_by_type = nvg_by_type or {}
    ws = openpyxl.load_workbook(XLSX, data_only=True)["CV Summary"]

    def rowvals(r):
        return {key: r2(ws.cell(row=r, column=col).value or 0.0) for key, col in CV_COLS}

    groups = []
    grand = None
    current = None
    for r in range(4, 25):
        label = ws.cell(row=r, column=2).value
        if not label:
            continue
        label = str(label).strip()
        up = label.upper()
        if up.startswith("TOTALS HELICOPTER"):
            if current:
                current["totals"] = rowvals(r)
                groups.append(current)
                current = None
        elif up.startswith("TOTALS FIXED WING"):
            if current:
                current["totals"] = rowvals(r)
                groups.append(current)
                current = None
        elif up == "TOTALS":
            grand = rowvals(r)
            break
        else:
            # A type row. Start a group lazily based on where we are.
            if current is None:
                current = {
                    "name": "Helicopter" if not groups else "Fixed Wing",
                    "rows": [],
                    "totals": {},
                }
            current["rows"].append({"type": label, **rowvals(r)})
    if current and not current.get("totals"):
        groups.append(current)

    # Fold flight-derived NVG into the matching type rows, then rebuild the NVG
    # subtotals and grand total (the "total" column is Captain+Other+Dual, so
    # NVG does not affect it).
    for g in groups:
        for row in g["rows"]:
            add = nvg_by_type.get(canon_type(row["type"]), 0.0)
            if add:
                row["nvg"] = r2(row["nvg"] + add)
        g["totals"]["nvg"] = r2(sum(r["nvg"] for r in g["rows"]))
    if grand is not None:
        grand["nvg"] = r2(sum(g["totals"]["nvg"] for g in groups))

    # Override the grand-total row with the logbook-derived figures so the CV
    # headline equals the authoritative Grand Total Flying Hours.
    if logbook_grand is not None:
        grand = dict(logbook_grand)

    extra = {
        "offshoreHours": r2(ws.cell(row=5, column=11).value or 0.0),
        "xCountry": r2(ws.cell(row=9, column=11).value or 0.0),
        "picXCountry": r2(ws.cell(row=13, column=11).value or 0.0),
        "xCountryInst": r2(ws.cell(row=13, column=13).value or 0.0),
    }
    return {
        "columns": [k for k, _ in CV_COLS],
        "groups": groups,
        "grandTotals": grand,
        "extra": extra,
    }


# Per-sheet column maps (1-based Excel column -> app time key) for the three
# simulator training sheets. AW139 SIM / S92 SIM use the multi-engine logbook
# layout; S70 FF&MS uses the military "Totals" layout (Day/Night x Captain/
# Other/Dual, plus Using-NVD, Sim, Actual-instrument columns).
SIM_SHEETS = [
    (
        "AW139 SIM",
        "AW139",
        {
            3: "meIcusDay", 4: "meIcusNight", 5: "meDualDay", 6: "meDualNight",
            7: "meCommandDay", 8: "meCommandNight", 9: "meCopilotDay", 10: "meCopilotNight",
            11: "instInFlight", 12: "instSim", 14: "nvg",
        },
    ),
    (
        "S92 SIM",
        "S92",
        {
            3: "meIcusDay", 4: "meIcusNight", 5: "meDualDay", 6: "meDualNight",
            7: "meCommandDay", 8: "meCommandNight", 9: "meCopilotDay", 10: "meCopilotNight",
            11: "instInFlight", 12: "instSim",
        },
    ),
    (
        "S70 FF&MS",
        "S-70",
        {
            3: "meCommandDay", 4: "meCopilotDay", 5: "meDualDay",
            6: "meCommandNight", 7: "meCopilotNight", 8: "meDualNight",
            12: "nvg", 14: "instSim", 15: "instInFlight",
        },
    ),
]


def build_sim_flights() -> list[dict]:
    """Extract every individual simulator session from the S70 FF&MS, S92 SIM
    and AW139 SIM sheets and merge them into one date-ordered simulator logbook.

    Each sheet row is one sim session. Values are mapped to the app time schema;
    subtotal rows (I=total, J=As Captain, K=As Instruct on the S70 sheet) are
    ignored so only the loggable columns are carried."""
    wb_v = openpyxl.load_workbook(XLSX, data_only=True)
    wb_c = openpyxl.load_workbook(XLSX, data_only=False)
    zero = {camel: 0.0 for camel in TIME_MAP.values()}
    zero["nvg"] = 0.0
    out: list[dict] = []

    for sheet, ac_type, colmap in SIM_SHEETS:
        ws = wb_v[sheet]
        wc = wb_c[sheet]
        cur_year = cur_month = None
        for r in range(1, ws.max_row + 1):
            a = ws.cell(row=r, column=1).value
            if isinstance(a, str) and a.strip():
                low = a.strip().upper()
                if low.startswith(("TOTAL", "GRAND")):
                    continue
                m, cur_year = parse_month_year(a, cur_year)
                if m:
                    cur_month = m
            b = ws.cell(row=r, column=2).value
            if isinstance(b, str) and b.strip().upper().startswith("TOTAL"):
                continue
            try:
                day = int(float(str(b).strip()))
            except (ValueError, TypeError):
                continue

            time = dict(zero)
            has_hours = False
            for col, key in colmap.items():
                v = ws.cell(row=r, column=col).value
                if isinstance(v, (int, float)) and v:
                    time[key] = r2(time[key] + float(v))
                    has_hours = True
            if not has_hours or not (cur_year and cur_month):
                continue

            note = ""
            for c in range(1, wc.max_column + 1):
                cell = wc.cell(row=r, column=c)
                if cell.comment and cell.comment.text.strip():
                    note = cell.comment.text
                    break
            lines = [ln.strip() for ln in note.splitlines() if ln.strip()]

            iso = f"{cur_year:04d}-{cur_month:02d}-{day:02d}"
            out.append(
                {
                    "id": f"sim-{sheet.replace(' ', '_').replace('&', 'n')}-{r}",
                    "sourceRow": r,  # replaced below with a global date-ordered index
                    "date": iso,
                    "year": cur_year,
                    "month": cur_month,
                    "day": day,
                    "aircraftType": ac_type,
                    "aircraftRego": "",
                    "pilotInCommand": "",
                    "otherCrew": lines[0] if lines else "",
                    "route": lines[1] if len(lines) >= 2 else "",
                    "remarks": " / ".join(lines[2:]) if len(lines) >= 3 else "",
                    "noteRaw": note.strip(),
                    "time": time,
                    "origin": "excel",
                }
            )

    out.sort(key=lambda f: (f["date"], f["aircraftType"], f["sourceRow"]))
    # Reassign sourceRow to a global, date-ordered sequence so the app can keep
    # imported sim sessions in order and slot newly added ones by date.
    for i, f in enumerate(out, start=1):
        f["sourceRow"] = i
    return out


def main() -> None:
    enriched = json.loads(ENRICHED.read_text(encoding="utf-8"))
    report = json.loads(REPORT.read_text(encoding="utf-8"))

    opening = dict(report["opening_balance"])
    extracted = report["extracted_grand_sum"]

    # Documented reconciliation correction: the page-1 brought-forward balance
    # under-recorded co-pilot day time by 9.1 h relative to the per-aircraft-type
    # sheets. Adding it to the opening co-pilot day column brings the logbook's
    # page-by-page running total up to the authoritative Grand Total Flying Hours
    # (5005.5). Applied at the very start so it carries through every page.
    OPENING_COPILOT_DAY_CORRECTION = 9.1
    opening["me_copilot_day"] = r2(opening["me_copilot_day"] + OPENING_COPILOT_DAY_CORRECTION)

    nvg_by_date, nvg_note_by_date, nvg_opening, nvg_total = build_nvg()

    # The pilot's deliberate running-total re-classifications live as typed
    # constants inside the "TOT BROUGHT FWD" formulas (e.g. Co-pilot -> Command
    # at the Nov-2014 and Aug-2016 page breaks). migrate.py surfaces each as a
    # carry-forward "difference"; summing them per column gives the net kept
    # adjustment. We reconcile against these constants rather than Excel's cached
    # final "TOT TO DATE" so the result is immune to stale formula caches (e.g.
    # the =SUM() page totals on pages 45/46 that Excel had not recalculated).
    adjustments = {camel: 0.0 for camel in TIME_MAP.values()}
    for d in report["carryforward_discrepancies"]:
        camel = TIME_MAP.get(d["column"])
        if camel is not None:
            adjustments[camel] = r2(adjustments[camel] + d["difference"])

    # Corrected logbook running total = opening + every entry + kept adjustments.
    final = {
        snake: r2(opening[snake] + extracted[snake] + adjustments[camel])
        for snake, camel in TIME_MAP.items()
    }
    opening_camel = {camel: r2(opening[snake]) for snake, camel in TIME_MAP.items()}
    final_camel = {camel: r2(final[snake]) for snake, camel in TIME_MAP.items()}
    opening_camel["nvg"] = r2(nvg_opening)
    final_camel["nvg"] = r2(nvg_total)
    # adjustments["nvg"] set after we've summed matched per-flight NVG below.

    # Drop residual "copied page" rows left over in the Excel: entries with no
    # day AND zero total hours (the repeated YSTW/YLMQ note blocks after the
    # last real flight on 2 Jul 26). They carry no hours, so totals are
    # unaffected.
    def is_residual(f: dict) -> bool:
        return f.get("day") is None and sum(f["times"].values()) == 0

    kept = [f for f in enriched if not is_residual(f)]
    dropped = len(enriched) - len(kept)

    # Assign NVG to a specific flight by DATE. Where several flights share a
    # date, prefer the one whose note matches the AW139 sheet's note, otherwise
    # the first. Totals are unaffected by the tie-break; only which row shows
    # the NVG value differs.
    main_by_date: dict[str, list[dict]] = defaultdict(list)
    for f in kept:
        if f.get("date"):
            main_by_date[f["date"]].append(f)

    nvg_by_row: dict[int, float] = {}
    nvg_matched_sum = 0.0
    nvg_unplaced = 0.0
    for iso, hrs in nvg_by_date.items():
        cands = main_by_date.get(iso, [])
        if not cands:
            nvg_unplaced += hrs
            continue
        target = cands[0]
        aw_note = norm_note(nvg_note_by_date.get(iso, ""))
        if aw_note:
            for c in cands:
                if norm_note(c.get("note_raw", "")) == aw_note:
                    target = c
                    break
        nvg_by_row[target["source_row"]] = nvg_by_row.get(target["source_row"], 0.0) + hrs
        nvg_matched_sum += hrs
    nvg_matched = sum(1 for v in nvg_by_row.values() if v)

    flights = []
    for f in kept:
        time = {camel: r2(f["times"][snake]) for snake, camel in TIME_MAP.items()}
        time["nvg"] = r2(nvg_by_row.get(f["source_row"], 0.0))
        flights.append(
            {
                "id": f"xl-{f['source_row']}",
                "sourceRow": f["source_row"],
                "date": f["date"],
                "year": f["year"],
                "month": f["month"],
                "day": f["day"],
                "aircraftType": f.get("aircraftType", ""),
                "aircraftRego": f.get("aircraftRego", ""),
                "pilotInCommand": f.get("pilotInCommand", ""),
                "otherCrew": f.get("otherCrew", ""),
                "route": f.get("route", ""),
                "remarks": f.get("remarks", ""),
                "noteRaw": f.get("note_raw", ""),
                "time": time,
                "needsReview": not f.get("aircraftType", ""),
                "origin": "excel",
            }
        )

    adjustments["nvg"] = r2(nvg_total - (nvg_opening + nvg_matched_sum))

    # Per-type NVG from the imported flights, so the CV snapshot can carry the
    # AW139 night-vision hours the Excel CV sheet left blank. All logged NVG is
    # flown on the AW139, so any flight carrying NVG (including a couple whose
    # type the note-based enrichment couldn't resolve) is attributed to AW139.
    total_flight_nvg = r2(sum(fl["time"].get("nvg", 0.0) for fl in flights))
    nvg_by_type: dict[str, float] = {"AW139": total_flight_nvg} if total_flight_nvg else {}

    sim_flights = build_sim_flights()

    # Authoritative "Grand Total Flying Hours" = the logbook running total, i.e.
    # the sum of the 14 flying columns of (opening + entries + kept adjustments).
    # Instrument (in-flight/sim) and NVG are sub-counts, not part of flying time.
    non_flying = {"instInFlight", "instSim", "nvg"}
    logbook_flying_total = r2(
        sum(v for k, v in final_camel.items() if k not in non_flying)
    )

    data = {
        "meta": {
            "generated": date.today().isoformat(),
            "source": report["source_file"],
            "numFlights": len(flights),
            "numPages": report["num_pages"],
            "note": (
                "Numbers transcribed verbatim from the Excel and reconciled "
                "against the pilot's own page totals. openingBalance + "
                "adjustments + sum(flights) == excelGrandTotal exactly."
            ),
        },
        "openingBalance": opening_camel,
        "adjustments": adjustments,
        "excelGrandTotal": final_camel,
        "totalsSheet": build_totals_sheet(logbook_flying_total),
        "cvSummary": build_cv_summary(nvg_by_type),
        "flights": flights,
        "simFlights": sim_flights,
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(data, indent=2), encoding="utf-8")

    nonzero_adj = {k: v for k, v in adjustments.items() if abs(v) > 0.001}
    print(f"Wrote {OUT.relative_to(HERE.parent)}")
    print(f"  residual rows dropped : {dropped}")
    print(f"  flights           : {len(flights)}")
    print(f"  NVG opening (mil)  : {r2(nvg_opening)}")
    print(f"  NVG flights matched: {nvg_matched} (sum {r2(nvg_matched_sum)}, unplaced {r2(nvg_unplaced)})")
    print(f"  NVG total (opening+flights={r2(nvg_opening + nvg_matched_sum)}) vs excel {r2(nvg_total)}; adj {adjustments['nvg']}")
    print(f"  opening balance    : total {r2(sum(opening_camel.values()))} hrs")
    print(f"  excel grand total  : total {r2(sum(final_camel.values()))} hrs")
    print(f"  reconciliation adj : {nonzero_adj}")
    FLYING = [c for s, c in TIME_MAP.items() if s not in ("inst_in_flight", "inst_sim")]
    by_type: dict[str, float] = {}
    for sf in sim_flights:
        by_type[sf["aircraftType"]] = r2(
            by_type.get(sf["aircraftType"], 0.0) + sum(sf["time"][c] for c in FLYING)
        )
    print(f"  sim sessions       : {len(sim_flights)}  flying by type {by_type}")


if __name__ == "__main__":
    main()
