"""Check whether specific LOG BOOK flights already appear on the per-type
sheets (so we don't double-count when reconciling)."""

from datetime import datetime
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent
wb = openpyxl.load_workbook(HERE / "_tmp_read.xlsx", data_only=True)


def parse_my(v, last):
    if v is None:
        return last
    s = str(v).strip()
    for fmt in ("%b %y", "%b-%y", "%B %y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return last


def dump(sheet, year, month, ncols=16):
    ws = wb[sheet]
    my = None
    print(f"\n=== {sheet}: rows in {month:02d}/{year} ===")
    hit = False
    for r in range(4, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        my = parse_my(a, my)
        try:
            day = int(float(str(b).strip()))
        except (TypeError, ValueError):
            continue
        if my and my.year == year and my.month == month:
            vals = [ws.cell(row=r, column=c).value for c in range(3, ncols)]
            vals = [v for v in vals if v not in (None, 0)]
            # first comment on the row, if any
            cm = ""
            for c in range(1, ncols):
                cell = ws.cell(row=r, column=c)
                if cell.comment and cell.comment.text.strip():
                    cm = cell.comment.text.strip().splitlines()[0]
                    break
            print(f"   row{r} day {day}: {vals}   note='{cm}'")
            hit = True
    if not hit:
        print("   (no rows found)")


# Group B checks
dump("S92", 2009, 12)     # rows 104/105 (14th, 16th)
dump("S92", 2016, 7)      # row 974 (31st)
dump("AW139", 2026, 2)    # row 1831 (19th)
# FSI check
dump("S92 SIM", 2009, 10)
dump("AW139 SIM", 2009, 10)
